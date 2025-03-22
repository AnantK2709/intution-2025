from config import client
import uuid
import os
from openpyxl import Workbook, load_workbook



FEEDBACK_FILE = "feedback.xlsx"

def generate_prompt(technology, framework, audience):
    return f"""
You are a Change Management Expert at MSD Company, specializing in helping organizations smoothly transition to new technologies using structured frameworks.

A company is planning to adopt {technology}, and they want to follow the {framework} framework to ensure a structured and effective transition. Your task is to create a comprehensive, practical, and actionable guide tailored to the needs of {Audience}.

Guide Requirements:
Step-by-Step Implementation Plan

Provide a detailed, structured rollout plan following the {framework} methodology.
Each step should be numbered, easy to follow, and include clear instructions.
Offer real-world examples or case studies relevant to {technology} adoption.
Include best practices to ensure success.
Practical Resources

Use Markdown formatting for clarity.
Where applicable, provide:
Templates for planning, implementation, and feedback collection.
Checklists for teams to ensure smooth adoption.
Common Technical FAQs and Answers

Identify common technical questions employees may have about {technology}.
Provide clear, concise answers to address concerns.
Offer troubleshooting tips for common issues during adoption.
Engagement and Support Plan

Outline strategies to ensure employees are engaged and well-supported.
Include training recommendations (e.g., workshops, online courses, mentoring).
Suggest ways to measure success and gather feedback for continuous improvement.
Tone and Style:
Maintain a helpful, structured, and professional tone.
Ensure the guide is easy to understand for {audience}.
Prioritize practicality over theory—focus on real actions companies can take.
"""

def generate_adoption_guide(prompt):
    response = client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[{"role": "user", "content": prompt}],
        temperature=0.2,
    )
    return response.choices[0].message.content

def refine_prompt_with_feedback(original_prompt, combined_feedback):
    improvement_request = f"""
You are a prompt engineer. Here's a prompt that was used with GPT:

---
{original_prompt}
---

The user was not satisfied with the result and gave the following feedback:
{combined_feedback}

Please improve the original prompt based on this feedback. Only output the improved prompt, nothing else.
"""

    response = client.chat.completions.create(
        model="gpt-4o-mini",
        messages=[{"role": "user", "content": improvement_request}],
        temperature=0.3,
    )
    return response.choices[0].message.content

# === Excel Logic ===
def initialize_excel():
    if not os.path.exists(FEEDBACK_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Feedback"
        ws.append(["ID", "Original Prompt", "User Feedback", "Status"])
        wb.save(FEEDBACK_FILE)

def save_feedback_to_excel(original_prompt, feedback_text):
    initialize_excel()
    wb = load_workbook(FEEDBACK_FILE)
    ws = wb["Feedback"]
    ws.append([str(uuid.uuid4()), original_prompt.strip(), feedback_text.strip(), "new"])
    wb.save(FEEDBACK_FILE)

def get_feedback_batch(limit=5):
    wb = load_workbook(FEEDBACK_FILE)
    ws = wb["Feedback"]
    batch = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[3] == "new":
            batch.append(row)
        if len(batch) == limit:
            break
    return batch

def mark_feedback_as_processed(batch_ids):
    wb = load_workbook(FEEDBACK_FILE)
    ws = wb["Feedback"]
    for row in ws.iter_rows(min_row=2):
        if row[0].value in batch_ids:
            row[3].value = "processed"
    wb.save(FEEDBACK_FILE)

# === Handler for FastAPI ===
def run_strategy_workflow(technology, framework,audience, feedback=None):
    original_prompt = generate_prompt(technology, framework,audience)
    guide = generate_adoption_guide(original_prompt)

    response = {
        "original_prompt": original_prompt,
        "guide": guide,
        "improved_prompt": None,
        "improved_guide": None
    }

    if feedback:
        save_feedback_to_excel(original_prompt, feedback)
        batch = get_feedback_batch()
        if len(batch) >= 5:
            combined_feedback = "\n".join([f"- {entry[2]}" for entry in batch])
            improved_prompt = refine_prompt_with_feedback(original_prompt, combined_feedback)
            improved_guide = generate_adoption_guide(improved_prompt)
            mark_feedback_as_processed([entry[0] for entry in batch])

            response["improved_prompt"] = improved_prompt
            response["improved_guide"] = improved_guide

    return response

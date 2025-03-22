import openai
import uuid
import os
from openpyxl import Workbook, load_workbook

# === Config ===
openai.api_key = "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"
FEEDBACK_FILE = "feedback.xlsx"

# === Prompt Logic ===
def generate_prompt(technology, framework):
    return f"""
You are a change management expert helping companies adopt new technologies using structured frameworks.

A company wants to adopt a new technology: {technology}
They want to follow the {framework} framework.

Give a clear, step-by-step guide tailored to employees, using the {framework} model. Make it practical and actionable.
Use numbered steps and keep the tone helpful.
Also include a list of common technical FAQs regarding the technology and answers to help employees understand the change process.
"""

def generate_adoption_guide(prompt):
    response = openai.chat.completions.create(
        model="gpt-4o-mini",
        messages=[{"role": "user", "content": prompt}],
        temperature=0.2,
    )
    return response.choices[0].message.content

def refine_prompt_with_feedback(original_prompt, combined_feedback):
    improvement_request = f"""
You are a prompt engineer. Here's a prompt that was used with GPT:

---
{original_prompt}
---

The user was not satisfied with the result and gave the following feedback:
{combined_feedback}

Please improve the original prompt based on this feedback. Only output the improved prompt, nothing else.
"""

    response = openai.chat.completions.create(
        model="gpt-4o-mini",
        messages=[{"role": "user", "content": improvement_request}],
        temperature=0.3,
    )
    return response.choices[0].message.content

# === Excel Feedback Storage ===
def initialize_excel():
    if not os.path.exists(FEEDBACK_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Feedback"
        ws.append(["ID", "Original Prompt", "User Feedback", "Status"])
        wb.save(FEEDBACK_FILE)

def save_feedback_to_excel(original_prompt, feedback_text):
    initialize_excel()
    wb = load_workbook(FEEDBACK_FILE)
    ws = wb["Feedback"]
    ws.append([str(uuid.uuid4()), original_prompt.strip(), feedback_text.strip(), "new"])
    wb.save(FEEDBACK_FILE)

def get_feedback_batch(limit=5):
    wb = load_workbook(FEEDBACK_FILE)
    ws = wb["Feedback"]
    batch = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[3] == "new":
            batch.append(row)
        if len(batch) == limit:
            break
    return batch

def mark_feedback_as_processed(batch_ids):
    wb = load_workbook(FEEDBACK_FILE)
    ws = wb["Feedback"]
    for row in ws.iter_rows(min_row=2):
        if row[0].value in batch_ids:
            row[3].value = "processed"
    wb.save(FEEDBACK_FILE)

# === Main Assistant Function ===
def run_change_assistant(technology, framework):
    original_prompt = generate_prompt(technology, framework)
    guide = generate_adoption_guide(original_prompt)

    print("\n📘 Change Management Guide:\n")
    print(guide)

    satisfied = input("\nAre you happy with this output? (yes/no): ").strip().lower()

    if satisfied == "no":
        feedback = input("Please enter a short one-line feedback to improve it: ").strip()
        save_feedback_to_excel(original_prompt, feedback)
        print("📝 Feedback saved to Excel. Will regenerate after 5 entries.")

        batch = get_feedback_batch()
        if len(batch) >= 5:
            print("\n🔁 5 feedback entries received. Generating improved prompt...\n")
            combined_feedback = "\n".join([f"- {entry[2]}" for entry in batch])
            improved_prompt = refine_prompt_with_feedback(original_prompt, combined_feedback)
            mark_feedback_as_processed([entry[0] for entry in batch])
            print("\n🧠 Improved Prompt:\n", improved_prompt)
            improved_guide = generate_adoption_guide(improved_prompt)
            print("\n📘 Updated Change Management Guide:\n")
            print(improved_guide)
    else:
        print("✅ Awesome! Share this guide with your team.")
